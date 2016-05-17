from openpyxl import *

print("reading dummy file")
dum = load_workbook('dummy_file.xlsx')
sheet = dum.get_sheet_by_name("Sheet1")

names = sheet.rows

sheet.cell(row=1, column=4).value = "new event"
print(sheet.cell(row=1,column=4).value)

dum.save('dummy_file.xlsx')

#display data

r = ''
for i in range(1,len(names)):
	print("----------------")
	for j in range(0,len(names[i])):
		r = r + str(names[i][j].value) + '-'
	print(r)
	r = ''
print("************")