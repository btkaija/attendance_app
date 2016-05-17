from openpyxl import *

class CardReader:
	
	def __init__(self, filename = 'Pi_Kappa_Phi_Attendance.xlsx'):
		#default is the pi kapp attendance sheet
		self.f = filename
	
	def read_first_name(self):
		name = input("What is your first name?")
		return name
	
	def read_last_name(self):
		name = input("What is your last name?")
		return name

	# Will either return a nine diget ID number or 'invalid'
	def read_card(self, swipe):
		#normal swipe is formatted as ;XXXXXXXXX=XXXX?
		#faulty swipe %E?;XXXXXXXXX=XXXX? or ;E? or ;E?+E? or +E?
		if len(swipe) == 9:
			return swipe
		elif len(swipe) == 16:
			normal_swipe = swipe.split('=')[0].split(';')[1]
			if len(normal_swipe) != 9:
				normal_swipe = 'invalid'
			return normal_swipe
		elif len(swipe) == 19:
			errored_swipe = swipe.split(';')[1].split('=')[0]
			if len(errored_swipe) != 9:
				errored_swipe = 'invalid'
			return errored_swipe
		else:
			return 'invalid'
	
	def contains_ID(self, value):
		is_present = 0
		wb = load_workbook(self.f)
		ws = wb.active
		ids = (ws.columns)[0]
		for i in range(1,len(ids)):
			if(int(value) == ids[i].value):
				is_present = 1
		return bool(is_present == 1)	

	def get_first_name_from_ID(self, value):
		wb = load_workbook(self.f)
		ws = wb.active
		people = ws.rows
		for i in range(1, len(people)):
			if(people[i][0].value == int(value)):
				return people[i][1].value

	def get_last_name_from_ID(self, value):
		wb = load_workbook(self.f)
		ws = wb.active
		people = ws.rows
		for i in range(1, len(people)):
			if(people[i][0].value == int(value)):
				return people[i][2].value
	
	def add_ID_to_name(self, value, first_name, last_name):
		successful_add = 0
		wb = load_workbook(self.f)
		ws = wb.active
		people = ws.rows
		for i in range(1, len(people)):
			if(first_name == people[i][1].value and last_name == people[i][2].value):
				people[i][0].value = int(value)
				successful_add = 1
		wb.save(self.f)
		return bool(successful_add == 1)

	def add_person_to_doc(self, value, first_name, last_name):
		wb = load_workbook(self.f)
		ws = wb.active
		people = ws.rows
		num_people = len(people)
		ws.cell(row = num_people+1, column=1).value = int(value)
		ws.cell(row = num_people+1, column=2).value = first_name
		ws.cell(row = num_people+1, column=3).value = last_name
		wb.save(self.f)
		return True

	def contains_event_name(self, value):
		wb = load_workbook(self.f)
		ws = wb.active
		column_names = (ws.rows)[0]
		for i in range(0, len(column_names)):
			if(column_names[i].value == value):
				return i

		return -1

	def add_event(self, value):
		wb = load_workbook(self.f)
		ws = wb.active
		num_columns = len((ws.rows)[0])
		rows = ws.rows
		ws.cell(row=1, column=num_columns+1).value = value
		for i in range(2, len(rows)+1):
			ws.cell(row = i, column = num_columns+1).value = 0
		wb.save(self.f)
		return num_columns+1

	def set_attended(self, value, event_index):
		wb = load_workbook(self.f)
		ws = wb.active
		people = ws.rows
		for i in range(1, len(people)):
			if(people[i][0].value == int(value)):
				ws.cell(row = i+1, column = event_index+1).value = 1
		wb.save(self.f)

