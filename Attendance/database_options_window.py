from tkinter import *
import tkinter as tk
import os
from openpyxl import *
class DatabaseOptionsWindow:

	def __init__(self):
		main_window = Tk()
		main_window.title("Export Options")
		#main_window.geometry('700x400+100+100')

		options_frame = Frame(main_window)
		options_frame.grid(column=0, row=0)
		self.prompt = Label(options_frame, text='What data would you like to export?',font = ("Corbel", "20", "bold"))
		self.prompt.grid(row = 0, column = 0, columnspan = 2)
		self.bro_check = Checkbutton(options_frame, text = 'Brother Data', command = self.check_toggled,font = ("Corbel", "20", "bold"))
		self.bro_check.grid(row = 1, column = 0)
		self.rush_check = Checkbutton(options_frame, text = 'Rush Data', command = self.check_toggled,font = ("Corbel", "20", "bold"))
		self.rush_check.grid(row = 1, column = 1)
		self.name_field = Entry(options_frame, width = 20 ,font = ("Corbel", "20", "bold"))
		self.name_field.grid(row = 2, column = 0, columnspan = 2)
		self.export_button = Button(options_frame, text = 'Export', command = self.export ,font = ("Corbel", "20", "bold"))
		self.export_button.grid(row = 4, column = 0, columnspan = 2)
		self.help = Label(options_frame, text = '(It will export to the current folder as a "_exported.xslx" file)')
		self.help.grid(row = 3, column = 0, columnspan = 2)
		self.error = Label(options_frame, text = "You cannot export with the file open!")

		self.broIsChecked = True
		self.bro_check.select()

		main_window.mainloop()

	def check_toggled(self):
		if(self.broIsChecked):
			self.broIsChecked = False
			self.bro_check.deselect()
			self.rush_check.select()
		else:
			self.broIsChecked = True
			self.bro_check.select()
			self.rush_check.deselect()
	def export(self):
		file_names = list(os.listdir('data/'))
		remove_items = []
		start_column = 0
		if(self.broIsChecked):
			start_column = 3
			master_list = open('IDs.dat')
			for i in range(len(file_names)):
				if 'Rushes' in file_names[i]:
					remove_items.append(file_names[i])
		else:
			start_column = 6
			master_list = open('Rushes.dat')
			for i in range(len(file_names)):
				if 'Rushes' not in file_names[i]:
					remove_items.append(file_names[i])
		for i in range(len(remove_items)):
			file_names.remove(remove_items[i])
		all_names = master_list.readlines()
		master_list.close()

		#start putting data in workbook

		wb = Workbook()
		ws = wb.active
		#intial data
		ws.cell(row = 1, column = 1).value = 'ID Number'
		ws.cell(row = 1, column = 2).value = 'Name'
		if(not self.broIsChecked):
			ws.cell(row = 1, column = 3).value = 'Email'
			ws.cell(row = 1, column = 4).value = 'Phone'
			ws.cell(row = 1, column = 5).value = 'Grad Year'
		#put event names in
		for i in range(start_column, len(file_names)+start_column):
			ws.cell(row = 1, column = i).value = file_names[i-start_column][:len(file_names[i-start_column])-4]
		#put all_names in
		for i in range(len(all_names)):
			personal_data = all_names[i].split('_')
			for j in range(len(personal_data)):
				ws.cell(row = 2+i, column = 1+j).value = personal_data[j]
		#put individual events in
		for x in range(len(file_names)):
			file_name = open('data/'+file_names[x])
			lines = file_name.readlines()

			for i in range(len(all_names)):
				val = 0
				for j in range(len(lines)):
					id_num = lines[j].split('_')[0]
					if(id_num == all_names[i].split('_')[0]):
						val = 1
						break
				ws.cell(row = 2+i, column = start_column+x).value = val
			file_name.close()
		#close
		#wb.save(self.name_field.get()+'_exported.xlsx')
		try:
			wb.save(self.name_field.get()+'_exported.xlsx')
			self.error.grid_remove()
		except PermissionError:
			self.error.grid(row= 5, column = 0, columnspan = 2)